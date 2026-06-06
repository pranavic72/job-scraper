# ============================================================
#  sheets.py — Excel output using openpyxl
# ============================================================

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = ["Company", "Role", "Link", "Status", "Date Found", "Location"]
EXCEL_FILE = "linkedin_jobs.xlsx"

STATUS_OPTIONS = '"Submitted,Rejected,Interviewing"'


def get_existing_links(ws) -> set:
    """Read all existing job links from column C to avoid duplicates."""
    links = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:  # Link is column C (index 2)
            links.add(row[2])
    return links


def add_status_dropdown(ws, start_row: int, end_row: int):
    """Add dropdown validation to the Status column (D) for given rows."""
    dv = DataValidation(
        type="list",
        formula1=STATUS_OPTIONS,
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = f"D{start_row}:D{end_row}"
    ws.add_data_validation(dv)


def style_header_row(ws):
    """Apply formatting to the header row."""
    header_fill = PatternFill(start_color="0A66C2", end_color="0A66C2", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 30  # Company
    ws.column_dimensions["B"].width = 40  # Role
    ws.column_dimensions["C"].width = 60  # Link
    ws.column_dimensions["D"].width = 18  # Status
    ws.column_dimensions["E"].width = 15  # Date Found
    ws.column_dimensions["F"].width = 25  # Location
    ws.row_dimensions[1].height = 20


def write_jobs_to_excel(jobs, filepath: str = EXCEL_FILE):
    """
    Write jobs to Excel file.
    Creates the file if it doesn't exist.
    Skips duplicates based on job link.
    Returns count of new jobs written.
    """
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        existing_links = get_existing_links(ws)
        print(f"📂  Opened existing file: {filepath}")
        print(f"    {len(existing_links)} existing job(s) found — duplicates will be skipped.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs"
        style_header_row(ws)
        existing_links = set()
        print(f"📄  Created new file: {filepath}")

    first_new_row = ws.max_row + 1
    new_count = 0

    for job in jobs:
        if job.link in existing_links:
            continue

        ws.append([
            job.company,    # A - Company
            job.role,       # B - Role
            job.link,       # C - Link
            "Update",             # D - Status (blank, user fills in)
            job.date_found, # E - Date Found
            job.location,   # F - Location
        ])

        # Make the link clickable
        row = ws.max_row
        link_cell = ws.cell(row=row, column=3)
        link_cell.hyperlink = job.link
        link_cell.font = Font(color="0A66C2", underline="single")

        existing_links.add(job.link)
        new_count += 1

    # Add status dropdown to all new rows
    last_row = ws.max_row
    if new_count > 0:
        add_status_dropdown(ws, first_new_row, last_row)

    wb.save(filepath)
    print(f"✅  {new_count} new job(s) written to {filepath}")
    return new_count
